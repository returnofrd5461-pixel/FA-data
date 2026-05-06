"""
build_data.py
raw/ 폴더 (재귀) Excel을 읽어 data/data.json 을 **월 단위 머지** 한다.

입력 (raw/**/*.xlsx 재귀 검색):
  - 손생보합산_*.xlsx → D     (FA·월 단위 통산유지율)
  - 통산유지율_*.xlsx → LOST  (실효·해지 건별)
  - 건별실적_*.xlsx   → PERF  (건당 실적 행 단위)

머지 원칙:
  - 기존 data.json 의 (FA, 월) 단위로 새 데이터만 덮어씀. 다른 월/기간은 보존.
  - PERF 의 totals/insurers 는 월 머지 후 자동 재계산/누적.
  - TARGET, FEEDBACK 키는 손대지 않음.

일시납 자동 정제 (PERF 행 단위):
  룰 A — 상품명/보험종목에 '연금' 포함 AND 영수보험료 ≥ 1천만 AND 납입회차=1 AND 영수유형='초회'
  룰 B — 영수보험료 ≥ 5천만 AND 납입회차=1 (안전망)
  제외 행은 logs/excluded_lump_sum_YYYYMMDD.log 에 기록.

사용법:
  python scripts/build_data.py            # 머지·쓰기
  python scripts/build_data.py --dry-run  # 머지만 시뮬레이션, 파일 쓰기 X
"""

from __future__ import annotations

import argparse
import datetime
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 이 설치되지 않았습니다: pip install openpyxl")

ROOT = Path(__file__).parent.parent
RAW = ROOT / "raw"
OUT = ROOT / "data" / "data.json"
LOG_DIR = ROOT / "logs"

Q4_MONTHS = ["10", "11", "12"]
Q1_MONTHS = ["01", "02", "03"]

# DB 배정 산정에서 무조건 제외할 FA.
# manual.json 에 db 값이 있어도 강제로 0 으로 덮어씀.
DB_EXCLUDE = {"신지원", "정민욱", "이연식", "김성한"}


# ── helpers ───────────────────────────────────────────────────────────────────

def find_files(pattern: str) -> list[Path]:
    """raw/ 재귀 검색 (직속 + 하위폴더). Excel 락파일(~$...) 제외."""
    if not RAW.exists():
        return []
    matches = sorted(set(RAW.rglob(pattern)))
    return [p for p in matches if not p.name.startswith("~$")]


def num(v) -> int:
    """콤마/공백 포함 문자열을 정수로. 변환 실패 시 0."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except (ValueError, OverflowError):
            return 0
    s = str(v).replace(",", "").strip()
    if not s:
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def find_header_row(rows, *labels) -> int:
    """주어진 라벨 중 하나가 포함된 첫 행을 헤더로 간주. 없으면 0."""
    label_set = set(labels)
    for i, r in enumerate(rows):
        if not r:
            continue
        cells = {str(c).strip() for c in r if c is not None}
        if cells & label_set:
            return i
    return 0


# ── 손생보합산 → D fragment ───────────────────────────────────────────────────

def build_D_fragment(path: Path) -> dict:
    """
    예상 컬럼: FA명, 팀, 월, 총계약건, 정상, 실효, 해지, 총보험료, 25회차유지율
    반환: {name: {name, team, months: {mm: {...}}}}
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}

    hdr_idx = find_header_row(rows, "FA명")
    headers = [str(h).strip() if h is not None else "" for h in rows[hdr_idx]]

    result: dict = {}
    for row in rows[hdr_idx + 1:]:
        if not row or all(c is None for c in row):
            continue
        rec = dict(zip(headers, row))
        name = str(rec.get("FA명", "") or "").strip()
        team = str(rec.get("팀", "") or "").strip()
        month = str(rec.get("월", "") or "").strip().zfill(2)
        if not name or not month:
            continue

        if name not in result:
            result[name] = {"name": name, "team": team, "months": {}}
        if team:
            result[name]["team"] = team

        result[name]["months"][month] = {
            "team":            team,
            "rate_25":         float(rec.get("25회차유지율") or 0),
            "total_contracts": num(rec.get("총계약건")),
            "normal":          num(rec.get("정상")),
            "lapsed":          num(rec.get("실효")),
            "cancelled":       num(rec.get("해지")),
            "total_prem":      float(rec.get("총보험료") or 0),
        }
    return result


# ── 통산유지율 → LOST fragment ────────────────────────────────────────────────

def build_LOST_fragment(path: Path) -> dict:
    """
    예상 컬럼: FA명, 기간, 보험사, 상품명, 계약자, 초회보험료, 현재상태,
              해지일, 계약일, 납입회차
    반환: {name: {period: [entry, ...]}}
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}

    hdr_idx = find_header_row(rows, "FA명")
    headers = [str(h).strip() if h is not None else "" for h in rows[hdr_idx]]

    result: dict = {}
    for row in rows[hdr_idx + 1:]:
        if not row or all(c is None for c in row):
            continue
        rec = dict(zip(headers, row))
        name = str(rec.get("FA명", "") or "").strip()
        period = str(rec.get("기간", "") or "").strip()
        if not name or not period:
            continue
        entry = {
            "insurer":     str(rec.get("보험사", "") or ""),
            "product":     str(rec.get("상품명", "") or ""),
            "holder":      str(rec.get("계약자", "") or ""),
            "first_perf":  num(rec.get("초회보험료")),
            "curr_status": str(rec.get("현재상태", "") or ""),
            "cancel_date": str(rec.get("해지일", "") or ""),
            "start_date":  str(rec.get("계약일", "") or ""),
            "paid_round":  num(rec.get("납입회차")),
        }
        result.setdefault(name, {}).setdefault(period, []).append(entry)
    return result


# ── 건별실적 → PERF fragment (행 단위 granular) ──────────────────────────────

def _extract_month(rec: dict) -> str:
    """납입월도(YYYYMM) → MM, 없으면 영수/환급일(YYYY-MM-DD) → MM."""
    pm = str(rec.get("납입월도", "") or "").strip()
    if pm and len(pm) >= 6 and pm[-2:].isdigit():
        return pm[-2:].zfill(2)
    rd = str(rec.get("영수/환급일", "") or "").strip()
    if len(rd) >= 7 and rd[4] == "-":
        return rd[5:7]
    cd = str(rec.get("확정일", "") or "").strip()
    if len(cd) >= 7 and cd[4] == "-":
        return cd[5:7]
    return ""


def _check_lump_sum(rec: dict) -> str | None:
    """일시납 룰 검사. 매치 시 사유 코드 반환, 아니면 None."""
    prem = num(rec.get("영수/환급보험료"))
    paid = num(rec.get("납입회차"))
    ptype = str(rec.get("영수유형", "") or "").strip()
    prod = str(rec.get("상품명", "") or "")
    sect = str(rec.get("보험종목", "") or "")

    # 룰 A: 연금 키워드 + 1천만+ + 1회차 + 초회
    if ("연금" in prod or "연금" in sect) and prem >= 10_000_000 and paid == 1 and ptype == "초회":
        return "A_연금일시납"
    # 룰 B: 5천만+ + 1회차
    if prem >= 50_000_000 and paid == 1:
        return "B_고액안전망"
    return None


def build_PERF_fragment(path: Path, exclude_log: list) -> dict:
    """
    행 단위 granular 데이터를 (FA, 월)로 집계.
    반환: {name: {name, months: {mm: {...}}, _insurers: {보험사: count}}}
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}

    hdr_idx = find_header_row(rows, "SUNAB_PK", "모집자명")
    headers = [str(h).strip() if h is not None else "" for h in rows[hdr_idx]]

    result: dict = {}
    for row in rows[hdr_idx + 1:]:
        if not row or all(c is None for c in row):
            continue
        rec = dict(zip(headers, row))
        name = str(rec.get("모집자명") or rec.get("담당자") or "").strip()
        if not name:
            continue
        month = _extract_month(rec)
        if not month:
            continue

        # 일시납 검사 → 제외
        excl = _check_lump_sum(rec)
        if excl:
            exclude_log.append({
                "fa":      name,
                "holder":  str(rec.get("계약자") or "").strip(),
                "prem":    num(rec.get("영수/환급보험료")),
                "product": str(rec.get("상품명") or "").strip(),
                "reason":  excl,
            })
            continue

        # 집계 버킷
        result.setdefault(name, {"name": name, "months": {}, "_insurers": {}})
        m_bucket = result[name]["months"].setdefault(month, {
            "cnt": 0, "prem": 0, "perf": 0, "hwan": 0,
            "life": 0, "nonlife": 0,
            "status":   {"정상": 0, "유예": 0, "해지": 0, "실효": 0},
            "products": {},
        })

        m_bucket["cnt"]  += 1
        m_bucket["prem"] += num(rec.get("월보험료"))
        m_bucket["perf"] += num(rec.get("인정실적"))
        m_bucket["hwan"] += num(rec.get("환산월초"))

        sect = str(rec.get("보험종목") or "")
        if "생명보험" in sect:
            m_bucket["life"] += 1
        else:
            m_bucket["nonlife"] += 1

        cs = str(rec.get("계약상태") or "").strip()
        if cs:
            m_bucket["status"][cs] = m_bucket["status"].get(cs, 0) + 1

        pg = str(rec.get("상품군") or "").strip()
        if pg:
            m_bucket["products"][pg] = m_bucket["products"].get(pg, 0) + 1

        ins = str(rec.get("보험회사") or "").strip()
        if ins:
            result[name]["_insurers"][ins] = result[name]["_insurers"].get(ins, 0) + 1

    return result


# ── merge ────────────────────────────────────────────────────────────────────

def merge_D(existing: dict, fragment: dict) -> dict:
    """D[name].months[month] 단위 덮어쓰기. 다른 월 보존."""
    out = json.loads(json.dumps(existing or {}))
    for name, frag in fragment.items():
        if name not in out:
            out[name] = {"name": name, "team": frag.get("team", ""), "status": "", "months": {}}
        if frag.get("team"):
            out[name]["team"] = frag["team"]
        out[name].setdefault("months", {})
        for m, v in frag.get("months", {}).items():
            out[name]["months"][m] = v
    return out


def merge_LOST(existing: dict, fragment: dict) -> dict:
    """LOST[name][period] 단위 덮어쓰기. 다른 기간 보존."""
    out = json.loads(json.dumps(existing or {}))
    for name, periods in fragment.items():
        out.setdefault(name, {})
        for period, entries in periods.items():
            out[name][period] = entries
    return out


def merge_PERF(existing: dict, fragment: dict, D_merged: dict) -> dict:
    """
    PERF[name].months[month] 단위 덮어쓰기 + insurers 추가 누적 + totals 재계산.
    insurers는 기존 누적 + 새 월 contribution (재실행 시 중복 누적 주의).
    """
    out = json.loads(json.dumps(existing or {}))
    for name, frag in fragment.items():
        if name not in out:
            team = (D_merged.get(name) or {}).get("team", "기타")
            out[name] = {
                "name":     name,
                "team":     team,
                "status":   "FA",       # 신규 FA 기본값. 기존 FA는 보존됨.
                "months":   {},
                "totals":   {},
                "insurers": {},
            }
        # team이 비어있으면 D에서 보강
        if not out[name].get("team"):
            d_team = (D_merged.get(name) or {}).get("team")
            if d_team:
                out[name]["team"] = d_team
        # 월 단위 머지
        out[name].setdefault("months", {})
        for m, v in frag.get("months", {}).items():
            out[name]["months"][m] = v
        # insurers 누적
        out[name].setdefault("insurers", {})
        for ins, c in frag.get("_insurers", {}).items():
            out[name]["insurers"][ins] = out[name]["insurers"].get(ins, 0) + c

    # totals 전체 재계산 (모든 FA — 기존 데이터 totals 정합성 유지)
    for name in out:
        recompute_totals(out[name])
    return out


def ensure_manual_only_perf(PERF: dict, TARGET: dict, FEEDBACK: dict, D: dict) -> list:
    """
    manual.json 으로만 등장한 FA (TARGET/FEEDBACK 에는 있지만 PERF/D 에는 없음)
    에게 빈 PERF entry 자동 생성. 카드 렌더 invariant 유지를 위함.

    team 은 '미배정' 디폴트. 추후 손생보합산(D) 데이터가 들어오면
    sync_team_from_D() 가 갱신.

    반환: 새로 추가한 FA 이름 리스트.
    """
    manual_names = set(TARGET.keys()) | set(FEEDBACK.keys())
    existing_names = set(PERF.keys()) | set(D.keys())
    new_names = sorted(manual_names - existing_names)

    for name in new_names:
        PERF[name] = {
            "name":     name,
            "team":     "미배정",
            "status":   "FA",
            "months":   {},
            "totals": {
                "cnt": 0, "prem": 0, "perf": 0, "hwan": 0,
                "life": 0, "nonlife": 0, "lost": 0, "delay": 0,
                "avg_perf": 0, "life_ratio": 0, "lost_rate": 0,
                "q4_cnt": 0, "q1_cnt": 0,
                "q4_perf": 0, "q1_perf": 0,
                "growth": None,
                "top_products": {},
            },
            "insurers": {},
        }
    return new_names


def sync_team_from_D(PERF: dict, D: dict) -> int:
    """team 이 비어있거나 '미배정'인 PERF entry 를 D 의 team 으로 보강."""
    fixed = 0
    for name, p in PERF.items():
        cur = p.get("team")
        if cur and cur != "미배정":
            continue
        d_team = (D.get(name) or {}).get("team")
        if d_team:
            p["team"] = d_team
            fixed += 1
    return fixed


def merge_manual(TARGET: dict, FEEDBACK: dict, path: Path) -> dict:
    """
    manual.json 한 개 파일을 TARGET / FEEDBACK 에 머지.
    - TARGET[FA][month] = {db, act, goal} — goal 은 기존 값 보존
    - FEEDBACK[FA][month] = {done, hold}
    - DB_EXCLUDE FA 는 db 값 강제 0
    - 기존에 없는 FA 는 자동 생성
    반환: 통계 dict (month, target_count, feedback_count, db_excluded)
    """
    data = json.loads(path.read_text(encoding="utf-8"))
    month = str(data.get("month", "")).strip().zfill(2)
    if not month:
        print(f"[WARN] {path}: 'month' 필드 없음 — 스킵")
        return {"path": str(path), "month": "", "target_count": 0,
                "feedback_count": 0, "db_excluded": 0}

    target_in   = data.get("target")   or {}
    feedback_in = data.get("feedback") or {}

    db_excluded_count = 0
    for name, vals in target_in.items():
        if not isinstance(vals, dict):
            continue
        db  = num(vals.get("db"))
        act = num(vals.get("act"))
        if name in DB_EXCLUDE:
            db = 0
            db_excluded_count += 1
        existing_month = (TARGET.get(name) or {}).get(month) or {}
        goal = existing_month.get("goal", 0)
        TARGET.setdefault(name, {})
        TARGET[name][month] = {"db": db, "act": act, "goal": goal}

    for name, vals in feedback_in.items():
        if not isinstance(vals, dict):
            continue
        done = num(vals.get("done"))
        hold = num(vals.get("hold"))
        FEEDBACK.setdefault(name, {})
        FEEDBACK[name][month] = {"done": done, "hold": hold}

    return {
        "path": str(path),
        "month": month,
        "target_count": len(target_in),
        "feedback_count": len(feedback_in),
        "db_excluded": db_excluded_count,
    }


def recompute_totals(entry: dict) -> None:
    """PERF[name] 의 totals 를 months 데이터로부터 새로 계산."""
    months = entry.get("months", {}) or {}

    cnt     = sum(m.get("cnt", 0)     for m in months.values())
    prem    = sum(m.get("prem", 0)    for m in months.values())
    perf    = sum(m.get("perf", 0)    for m in months.values())
    hwan    = sum(m.get("hwan", 0)    for m in months.values())
    life    = sum(m.get("life", 0)    for m in months.values())
    nonlife = sum(m.get("nonlife", 0) for m in months.values())

    lost = sum(
        (m.get("status", {}).get("실효", 0) + m.get("status", {}).get("해지", 0))
        for m in months.values()
    )
    delay = sum(m.get("status", {}).get("유예", 0) for m in months.values())

    avg_perf   = round(perf / cnt) if cnt > 0 else 0
    life_ratio = round(life / (life + nonlife) * 100, 1) if (life + nonlife) > 0 else 0
    lost_rate  = round(lost / cnt * 100, 1) if cnt > 0 else 0

    q4_cnt  = sum(months.get(m, {}).get("cnt", 0)  for m in Q4_MONTHS)
    q1_cnt  = sum(months.get(m, {}).get("cnt", 0)  for m in Q1_MONTHS)
    q4_perf = sum(months.get(m, {}).get("perf", 0) for m in Q4_MONTHS)
    q1_perf = sum(months.get(m, {}).get("perf", 0) for m in Q1_MONTHS)
    growth  = round((q1_perf - q4_perf) / q4_perf * 100) if q4_perf > 0 else None

    top_products: dict = {}
    for m_data in months.values():
        for k, v in (m_data.get("products") or {}).items():
            top_products[k] = top_products.get(k, 0) + v
    top_products = dict(sorted(top_products.items(), key=lambda x: -x[1]))

    entry["totals"] = {
        "cnt": cnt, "prem": prem, "perf": perf, "hwan": hwan,
        "life": life, "nonlife": nonlife, "lost": lost, "delay": delay,
        "avg_perf": avg_perf, "life_ratio": life_ratio, "lost_rate": lost_rate,
        "q4_cnt": q4_cnt, "q1_cnt": q1_cnt,
        "q4_perf": q4_perf, "q1_perf": q1_perf,
        "growth": growth,
        "top_products": top_products,
    }


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="data/data.json 월 단위 머지 빌더")
    parser.add_argument("--dry-run", action="store_true",
                        help="머지·정제만 시뮬레이션, 파일 쓰기 안 함")
    args = parser.parse_args()

    print("▶ 파일 탐색 중 (raw/ 재귀)...")
    f_d      = find_files("손생보합산*.xlsx")
    f_lost   = find_files("통산유지율*.xlsx")
    f_perf   = find_files("건별실적*.xlsx")
    f_manual = find_files("manual.json")

    for label, files in [("D", f_d), ("LOST", f_lost), ("PERF", f_perf),
                         ("Manual", f_manual)]:
        if files:
            print(f"  {label:<6} ← {[p.name for p in files]}")
        else:
            print(f"  {label:<6} (입력 파일 없음 — 기존 데이터 유지)")

    print("\n▶ 기존 data.json 로드...")
    existing: dict = {}
    if OUT.exists():
        try:
            existing = json.loads(OUT.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"[WARN] data.json 로드 실패: {e} — 빈 상태로 시작")
            existing = {}
    existing.setdefault("D", {})
    existing.setdefault("LOST", {})
    existing.setdefault("PERF", {})
    existing.setdefault("TARGET", {})
    existing.setdefault("FEEDBACK", {})

    print("\n▶ 머지 진행...")
    D_merged        = existing["D"]
    LOST_merged     = existing["LOST"]
    PERF_merged     = existing["PERF"]
    TARGET_merged   = existing["TARGET"]
    FEEDBACK_merged = existing["FEEDBACK"]
    exclude_log: list = []

    for path in f_d:
        D_merged = merge_D(D_merged, build_D_fragment(path))
    for path in f_lost:
        LOST_merged = merge_LOST(LOST_merged, build_LOST_fragment(path))
    for path in f_perf:
        frag = build_PERF_fragment(path, exclude_log)
        PERF_merged = merge_PERF(PERF_merged, frag, D_merged)

    # manual.json — TARGET / FEEDBACK 머지
    if f_manual:
        print(f"\n[manual] {len(f_manual)}개 파일 처리")
    for path in f_manual:
        stats = merge_manual(TARGET_merged, FEEDBACK_merged, path)
        rel = path.relative_to(ROOT).as_posix()
        m = stats["month"] or "??"
        print(f"  ← {rel}")
        print(f"    TARGET 갱신:   {m}월, FA {stats['target_count']}명 "
              f"(DB 강제 0: {stats['db_excluded']}명)")
        print(f"    FEEDBACK 갱신: {m}월, FA {stats['feedback_count']}명")

    # manual-only FA 빈 PERF entry 자동 생성 (카드 렌더 invariant)
    new_fa = ensure_manual_only_perf(PERF_merged, TARGET_merged, FEEDBACK_merged, D_merged)
    if new_fa:
        print(f"\n[manual-only FA] {len(new_fa)}명 빈 PERF entry 자동 생성: {', '.join(new_fa)}")
    # team='미배정' entry 가 D에 실제 team을 가지게 되면 자동 보강
    fixed = sync_team_from_D(PERF_merged, D_merged)
    if fixed:
        print(f"[team sync] {fixed}명의 PERF.team 을 D 데이터로 보강")

    # 일시납 요약
    print()
    if exclude_log:
        total = len(exclude_log)
        total_prem = sum(e["prem"] for e in exclude_log)
        unique_fa = len(set(e["fa"] for e in exclude_log))
        ra = sum(1 for e in exclude_log if e["reason"].startswith("A"))
        rb = sum(1 for e in exclude_log if e["reason"].startswith("B"))
        print(f"[일시납 제외] {total}건 / 총 {total_prem:,}원 / FA {unique_fa}명")
        print(f"  - 룰 A (연금 일시납): {ra}건")
        print(f"  - 룰 B (고액 안전망): {rb}건")
    else:
        print("[일시납 제외] 0건")

    # 데이터 월 union
    all_months: set = set()
    for e in PERF_merged.values():
        all_months.update((e.get("months") or {}).keys())
    for e in D_merged.values():
        all_months.update((e.get("months") or {}).keys())

    kst = datetime.timezone(datetime.timedelta(hours=9))
    now_kst = datetime.datetime.now(kst).replace(microsecond=0)

    out = {
        "_meta": {
            "lastUpdated": now_kst.isoformat(),
            "dataMonths":  sorted(all_months),
        },
        "D":        D_merged,
        "LOST":     LOST_merged,
        "PERF":     PERF_merged,
        "TARGET":   TARGET_merged,
        "FEEDBACK": FEEDBACK_merged,
    }

    if args.dry_run:
        print("\n[DRY-RUN] 파일 쓰기 생략. 머지 결과 요약:")
        print(f"  D={len(D_merged)}명  LOST={len(LOST_merged)}명  PERF={len(PERF_merged)}명")
        print(f"  데이터 월: {sorted(all_months)}")
        return

    # 일시납 로그 파일
    if exclude_log:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        log_path = LOG_DIR / f"excluded_lump_sum_{now_kst.strftime('%Y%m%d')}.log"
        with log_path.open("w", encoding="utf-8") as fh:
            fh.write("FA명\t계약자\t보험료\t상품명\t제외사유\n")
            for e in exclude_log:
                fh.write(f"{e['fa']}\t{e['holder']}\t{e['prem']}\t{e['product']}\t{e['reason']}\n")
        print(f"\n[일시납 로그] {log_path}")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, indent=2)

    print(f"\n✅ {OUT} 갱신 완료 ({now_kst.isoformat()})")
    print(f"   D={len(D_merged)}명  LOST={len(LOST_merged)}명  PERF={len(PERF_merged)}명")
    print(f"   데이터 월: {sorted(all_months)}")


if __name__ == "__main__":
    main()
